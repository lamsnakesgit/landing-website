import os
import json
import csv
from datetime import datetime
from pathlib import Path
import logging
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

def export_daily_leads(processed_leads, base_output_dir):
    """
    Экспорт собранных лидов с драфтами сообщений и офферами в папку с текущей датой:
    YYYY-MM-DD/
      ├── summary_leads.xlsx (Главная Excel таблица)
      ├── summary_leads.csv
      ├── summary_leads.json
      └── cards/ (Markdown карточки по каждому лиду)
    """
    today_str = datetime.now().strftime("%Y-%m-%d")
    target_folder = Path(base_output_dir) / today_str
    cards_folder = target_folder / "cards"

    target_folder.mkdir(parents=True, exist_ok=True)
    cards_folder.mkdir(parents=True, exist_ok=True)

    logger.info(f"📁 Сохраняем результаты сбора в папку: {target_folder}")

    # 1. Сохранение JSON
    json_path = target_folder / "summary_leads.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(processed_leads, f, ensure_ascii=False, indent=2)

    # 2. Сохранение CSV
    csv_path = target_folder / "summary_leads.csv"
    fieldnames = [
        "id", "date", "source", "query", "company", "contact_person",
        "contacts", "link", "niche", "draft_message", "offer_proposal", "details"
    ]
    with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for idx, lead in enumerate(processed_leads, 1):
            writer.writerow({
                "id": idx,
                "date": today_str,
                "source": lead.get("source", ""),
                "query": lead.get("query", ""),
                "company": lead.get("company", ""),
                "contact_person": lead.get("contact_person", ""),
                "contacts": lead.get("contacts", ""),
                "link": lead.get("link", ""),
                "niche": lead.get("niche", ""),
                "draft_message": lead.get("draft_message", ""),
                "offer_proposal": lead.get("offer_proposal", ""),
                "details": lead.get("details", "")
            })

    # 3. Сохранение стилизованной Таблицы Excel (.xlsx)
    xlsx_path = target_folder / "summary_leads.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Лиды {today_str}"
    ws.views.sheetView[0].showGridLines = True

    headers = [
        "№", "Источник", "Запрос", "Компания / Имя", "Контакты",
        "Ниша", "Драфт 1-го Сообщения", "Предложение (Оффер)", "Ссылка"
    ]

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    ws.append(headers)
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    for idx, lead in enumerate(processed_leads, 1):
        row_data = [
            idx,
            lead.get("source", ""),
            lead.get("query", ""),
            lead.get("company", ""),
            lead.get("contacts", ""),
            lead.get("niche", ""),
            lead.get("draft_message", ""),
            lead.get("offer_proposal", ""),
            lead.get("link", "")
        ]
        ws.append(row_data)
        row_idx = idx + 1

        for col_num in range(1, len(row_data) + 1):
            cell = ws.cell(row=row_idx, column=col_num)
            cell.border = thin_border
            if col_num in [1, 2, 3]:
                cell.alignment = center_align
            else:
                cell.alignment = left_align

    # Настройка ширины колонок
    col_widths = {
        1: 5,   # №
        2: 12,  # Источник
        3: 15,  # Запрос
        4: 25,  # Компания
        5: 30,  # Контакты
        6: 22,  # Ниша
        7: 45,  # Драфт 1-го сообщения
        8: 45,  # Предложение
        9: 30   # Ссылка
    }
    for col_num, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_num)].width = width

    wb.save(xlsx_path)

    # 4. Создание индивидуальных Markdown карточек для каждого лида
    for idx, lead in enumerate(processed_leads, 1):
        safe_company = "".join([c for c in lead.get("company", "lead") if c.isalnum() or c in (" ", "_", "-")]).strip()
        safe_company = safe_company.replace(" ", "_")[:30]
        card_name = f"{idx:02d}_{lead.get('source', 'lead')}_{safe_company}.md"
        card_path = cards_folder / card_name

        md_content = f"""# 📄 Карточка Лида №{idx}: {lead.get('company')}

- **Дата сбора**: {today_str}
- **Источник**: `{lead.get('source')}`
- **Поисковый запрос**: `{lead.get('query')}`
- **Направление/Ниша**: {lead.get('niche')}
- **Контактное лицо**: {lead.get('contact_person', 'Не указано')}
- **Контакты**: {lead.get('contacts')}
- **Ссылка**: [{lead.get('link')}]({lead.get('link')})

---

### 💬 Драфт 1-го Сообщения:
```text
{lead.get('draft_message')}
```

---

### 💡 Что предложить клиенту (Оффер):
```text
{lead.get('offer_proposal')}
```

---

### 🔍 Детали / Контекст запроса:
> {lead.get('details')}
"""
        with open(card_path, "w", encoding="utf-8") as f:
            f.write(md_content)

    logger.info(f"✨ Ежедневный отчёт успешно сформирован!")
    logger.info(f"📊 Всего лидов: {len(processed_leads)}")
    logger.info(f"📁 Папка дня: {target_folder}")

    return {
        "folder": str(target_folder),
        "xlsx": str(xlsx_path),
        "csv": str(csv_path),
        "json": str(json_path),
        "count": len(processed_leads)
    }
